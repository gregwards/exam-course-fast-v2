"""
Shared module for the framework's runner scripts.

Provides:
- Path constants
- Standards loaders (full + Part 1)
- Persona loader
- Design lock loader (from course-plan-exam.xlsx Course Narrative tab)
- Artifact loader (markdown / xlsx / csv)
- Finding writer (writes discriminator output to reviews/)

Stdlib + openpyxl only. Subject-agnostic; operates on file structure, not domain content.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Paths (resolved relative to the framework root)
# ---------------------------------------------------------------------------

# This module sits at scripts/_common.py; framework root is its parent.
FRAMEWORK_ROOT = Path(__file__).resolve().parent.parent

PATHS = {
    "standards": FRAMEWORK_ROOT / "standards" / "standards.md",
    "rize_lds": FRAMEWORK_ROOT / "standards" / "rize-lds.md",
    "personas": FRAMEWORK_ROOT / "personas",
    "prompts": FRAMEWORK_ROOT / "prompts",
    "templates": FRAMEWORK_ROOT / "templates",
    "research": FRAMEWORK_ROOT / "research",
}


# ---------------------------------------------------------------------------
# Course directory layout — per-course design folder
# ---------------------------------------------------------------------------

@dataclass
class CourseDirs:
    """Resolved per-course paths for one course's working folder."""
    root: Path
    design: Path
    output: Path
    reviews: Path
    data: Path

    @classmethod
    def from_root(cls, course_root: Path) -> "CourseDirs":
        course_root = Path(course_root).resolve()
        return cls(
            root=course_root,
            design=course_root / "design",
            output=course_root / "output",
            reviews=course_root / "reviews",
            data=course_root / "data",
        )

    def ensure(self) -> None:
        """Create the directory structure if it doesn't exist."""
        for p in [self.design, self.output, self.reviews, self.data]:
            p.mkdir(parents=True, exist_ok=True)
        for sub in ["persona-filter", "standards-audit", "subject-matter", "release"]:
            (self.reviews / sub).mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Standards loaders
# ---------------------------------------------------------------------------

def _read_text(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Required file not found: {path}")
    return path.read_text(encoding="utf-8")


def load_standards_full() -> str:
    """Return the combined Pillars + parent Rize LDS standards as one markdown blob.

    Order: parent first (general principles), then Pillars overrides/extensions.
    Useful for the standards-auditor when it needs the full corpus.
    """
    rize = _read_text(PATHS["rize_lds"])
    pillars = _read_text(PATHS["standards"])
    return f"# Parent — Rize LDS\n\n{rize}\n\n---\n\n# Pillars-specific\n\n{pillars}"


def _extract_section(markdown: str, heading_pattern: str) -> str:
    """Extract a markdown section by its heading.

    `heading_pattern` is matched as a regex against ## headers.
    Returns content from the matching ## header through the next ## header or EOF.
    """
    lines = markdown.splitlines()
    start: Optional[int] = None
    end: Optional[int] = None
    h2 = re.compile(r"^##\s+")
    for i, line in enumerate(lines):
        if start is None and re.match(rf"^##\s+{heading_pattern}", line):
            start = i
            continue
        if start is not None and h2.match(line):
            end = i
            break
    if start is None:
        return ""
    return "\n".join(lines[start:end] if end is not None else lines[start:])


def load_standards_part_1() -> str:
    """Return the foundational principles (Part 1) from both standards files.

    Includes LD-GEN-01 through LD-GEN-04 from parent (with LD-GEN-04 being the
    conflict-resolution rule), plus Pillars LD-PIL-01 from the Pillars layer.
    """
    rize = _read_text(PATHS["rize_lds"])
    pillars = _read_text(PATHS["standards"])

    # Parent: "## Part 1 — Foundational Principles" plus the conflict-resolution
    # section that lives at the top of the parent doc.
    rize_conflict = _extract_section(rize, r"Conflict Resolution")
    rize_part_1 = _extract_section(rize, r"Part 1\b")

    # Pillars: "## Part 1 — Design Principles" (LD-GEN-* notes + LD-PIL-01 Pillars-specific extension)
    pillars_part_1 = _extract_section(pillars, r"Part 1\b")

    sections = [
        "# Standards — Part 1 (Foundational Principles + Conflict Resolution)",
        "",
        "## Parent Rize LDS — Conflict Resolution",
        "",
        rize_conflict.replace("## Conflict Resolution", "").strip(),
        "",
        "## Parent Rize LDS — Foundational Principles",
        "",
        rize_part_1.replace("## Part 1 — Foundational Principles", "").strip(),
        "",
        "## Pillars-specific — Design Principles",
        "",
        pillars_part_1.replace("## Part 1 — Design Principles", "").strip(),
    ]
    return "\n".join(s for s in sections if s).strip()


# ---------------------------------------------------------------------------
# Persona loader
# ---------------------------------------------------------------------------

def load_persona(name: str = "online-skeptic") -> tuple[str, str]:
    """Load a persona card.

    Returns (path_str, content). Path is included so reviewer output can cite the source.
    """
    path = PATHS["personas"] / f"{name}.md"
    return str(path.relative_to(FRAMEWORK_ROOT)), _read_text(path)


# ---------------------------------------------------------------------------
# Design lock loader (Course Narrative from xlsx)
# ---------------------------------------------------------------------------

def load_design_lock(course_dirs: CourseDirs) -> str:
    """Extract the Course Narrative from the locked course-plan-exam.xlsx.

    Returns a markdown string suitable for injecting into reviewer prompts. Reads
    only the Course Narrative sheet — the LO Map and Assessment Map are not part
    of the design lock for the persona filter (the filter cares about opinion +
    acts + what-to-avoid + what-success-looks-like).
    """
    xlsx_path = course_dirs.design / "course-plan-exam.xlsx"
    if not xlsx_path.exists():
        return "(No design lock yet — course-plan-exam.xlsx not found in design/.)"

    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise ImportError(
            "openpyxl is required to read course-plan-exam.xlsx. "
            "Install with: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if "Course Narrative" not in wb.sheetnames:
        return "(No Course Narrative sheet found in course-plan-exam.xlsx.)"

    ws = wb["Course Narrative"]
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                value = str(cell).strip()
                if value:
                    lines.append(value)
                    break  # one cell per row in current template
    return "\n\n".join(lines).strip() or "(Course Narrative sheet is empty.)"


# ---------------------------------------------------------------------------
# Artifact loader (markdown / csv / xlsx)
# ---------------------------------------------------------------------------

def load_artifact(path: str | Path) -> str:
    """Load an artifact's content as a markdown-friendly string.

    - .md / .txt: returned as-is.
    - .csv: rendered as a markdown table.
    - .xlsx: each sheet rendered as a section with a markdown table.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Artifact not found: {p}")

    suffix = p.suffix.lower()
    if suffix in {".md", ".txt"}:
        return p.read_text(encoding="utf-8")

    if suffix == ".csv":
        with p.open(encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        return _rows_to_markdown_table(rows)

    if suffix == ".xlsx":
        return _xlsx_to_markdown(p)

    raise ValueError(f"Unsupported artifact type: {suffix} ({p})")


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    header = rows[0]
    body = rows[1:]
    sep = ["---"] * len(header)
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in body:
        # Pad short rows
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(str(c) for c in padded[: len(header)]) + " |")
    return "\n".join(lines)


def _xlsx_to_markdown(path: Path) -> str:
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise ImportError("openpyxl required to read xlsx artifacts") from e

    wb = openpyxl.load_workbook(str(path), data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [("" if c is None else str(c)) for c in row]
            # Trim trailing empties
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                rows.append(cells)
        if not rows:
            continue
        sections.append(f"## Sheet: {sheet_name}\n\n" + _rows_to_markdown_table(rows))
    return "\n\n".join(sections)


# ---------------------------------------------------------------------------
# Finding writer
# ---------------------------------------------------------------------------

def write_finding(
    course_dirs: CourseDirs,
    discriminator: str,
    artifact_id: str,
    reviewer_id: str,
    content: str,
) -> Path:
    """Write a discriminator finding to reviews/{discriminator}/.

    Filename: {artifact-id}-{reviewer-id}-review.md
    Returns the path written.
    """
    if discriminator not in {"persona-filter", "standards-audit", "subject-matter", "release"}:
        raise ValueError(f"Unknown discriminator: {discriminator}")
    course_dirs.ensure()
    safe_artifact = re.sub(r"[^A-Za-z0-9_-]+", "-", artifact_id).strip("-")
    safe_reviewer = re.sub(r"[^A-Za-z0-9_-]+", "-", reviewer_id).strip("-")
    fname = f"{safe_artifact}-{safe_reviewer}-review.md"
    out = course_dirs.reviews / discriminator / fname
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(content, encoding="utf-8")
    return out


# ---------------------------------------------------------------------------
# Misc helpers
# ---------------------------------------------------------------------------

def today() -> str:
    """Return today's date in YYYY-MM-DD."""
    return date.today().isoformat()
