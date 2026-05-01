#!/usr/bin/env python3
"""
D1 mechanical validation for the course plan xlsx.

Mirrors the spreadsheet's VALIDATION CHECKS formulas in Python so headless
agents have deterministic validation independent of any spreadsheet calc engine.
The spreadsheet's formulas remain the human-friendly source-of-truth definition;
this script is the agent-runtime mirror. Belt-and-suspenders.

Validation rules (from templates/source/course-plan-exam.xlsx VALIDATION CHECKS):
    Total = 1000 pts                       [LD-PIL-08, LD-GRD-01]
    Final exam 350-450 pts                 [LD-PIL-08, LD-PIL-21]
    Midterm 100-200 pts                    [LD-PIL-21]
    Act 1 ≤ 250 pts                        [LD-PIL-08]
    Max 5 units w/ secondary               [LD-PIL-09]
    No secondary > 15 pts                  [LD-PIL-10, LD-GRD-02]
    Midterm unit exam-only                 [LD-PIL-09]
    Final unit exam-only                   [LD-PIL-09]
    ≥ 9 exam-practice primaries            [LD-PIL-09]
    Skill benchmarked in Unit 6            [course-template requirement]

Plus standards-driven checks not in the spreadsheet:
    LO Map: ≤ 8 CLOs (LD-PIL-04)
    LO Map: 35-55 KLOs (LD-PIL-20)
    LO Map: every CLO references PLO as parent
    LO Map: every KLO references a CLO as parent
    Course Narrative: all 6 sections have content (Course Opinion, Act 1/2/3, Avoid, Success)

Usage:
    python3 scripts/check_course_plan.py --course /path/to/course-folder
    python3 scripts/check_course_plan.py --xlsx /path/to/course-plan-exam.xlsx
    [--report out.json]   write JSON report
    [--strict]            exit 1 if any FAIL (default: 0 always)

Output: prints a human-readable report to stdout. With --report, also writes JSON.
Exit: 0 always unless --strict is set.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Optional

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _common as common  # noqa: E402

# Assessment Map column indices (1-based, matching openpyxl)
COL = {
    "Unit": 1, "Act": 2, "Rize Requirements": 3, "Topic": 4,
    "CLOs Covered": 5, "KLOs Covered": 6, "Skill": 7,
    "Primary Assessment": 8, "Primary Pts": 9,
    "Secondary Assessment": 10, "Secondary Pts": 11,
    "Career Milestone": 12, "CM Pts": 13,
    "Unit Pts": 14, "Cumulative Pts": 15,
    "Primary Resource / Textbook Ch.": 16,
    "Unit Description": 17,
}

ASSESSMENT_DATA_ROWS = range(2, 16)  # rows 2-15 are units 1-14


@dataclass
class CheckResult:
    name: str
    standard: str
    severity: str  # PASS | FAIL | INFO
    message: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class Report:
    artifact: str
    results: list[CheckResult] = field(default_factory=list)

    def add(self, name: str, standard: str, severity: str, message: str) -> None:
        self.results.append(CheckResult(name=name, standard=standard, severity=severity, message=message))

    @property
    def fail_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "FAIL")

    @property
    def pass_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "PASS")

    @property
    def info_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "INFO")

    def to_dict(self) -> dict[str, Any]:
        return {
            "artifact": self.artifact,
            "summary": {
                "pass": self.pass_count,
                "fail": self.fail_count,
                "info": self.info_count,
            },
            "results": [r.to_dict() for r in self.results],
        }


def load_workbook(xlsx_path: Path):
    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise ImportError("openpyxl required: pip install openpyxl") from e
    return openpyxl.load_workbook(str(xlsx_path), data_only=True)


def _num(value: Any) -> Optional[float]:
    """Coerce a cell value to a number, returning None for blank/non-numeric."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _str(value: Any) -> str:
    """Coerce a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# Assessment Map checks (mirror the spreadsheet VALIDATION CHECKS)
# ---------------------------------------------------------------------------

def check_assessment_map(ws, report: Report) -> None:
    # Pull values once
    primary_pts = [_num(ws.cell(r, COL["Primary Pts"]).value) for r in ASSESSMENT_DATA_ROWS]
    secondary_assessment = [_str(ws.cell(r, COL["Secondary Assessment"]).value) for r in ASSESSMENT_DATA_ROWS]
    secondary_pts = [_num(ws.cell(r, COL["Secondary Pts"]).value) for r in ASSESSMENT_DATA_ROWS]
    cm_name = [_str(ws.cell(r, COL["Career Milestone"]).value) for r in ASSESSMENT_DATA_ROWS]
    unit_pts = [_num(ws.cell(r, COL["Unit Pts"]).value) for r in ASSESSMENT_DATA_ROWS]
    primary_assessment = [_str(ws.cell(r, COL["Primary Assessment"]).value) for r in ASSESSMENT_DATA_ROWS]
    skill = [_str(ws.cell(r, COL["Skill"]).value) for r in ASSESSMENT_DATA_ROWS]

    # Index helpers (1-based unit numbers map to 0-based list indices via index-1)
    def unit_pts_for(unit: int) -> Optional[float]:
        return unit_pts[unit - 1]

    # 1) Total = 1000 pts
    total = sum(p for p in unit_pts if p is not None)
    report.add(
        "Total = 1000 pts",
        "LD-PIL-08, LD-GRD-01",
        "PASS" if total == 1000 else "FAIL",
        f"Total: {total} pts" + ("" if total == 1000 else " (expected 1000)"),
    )

    # 2) Final exam 350-450 pts (Unit 14, Primary Pts)
    final_pts = primary_pts[13]  # unit 14 → index 13
    if final_pts is None:
        report.add("Final Exam 350-450 pts", "LD-PIL-08, LD-PIL-21", "FAIL", "Unit 14 Primary Pts is blank")
    elif 350 <= final_pts <= 450:
        report.add("Final Exam 350-450 pts", "LD-PIL-08, LD-PIL-21", "PASS", f"Unit 14: {final_pts} pts")
    else:
        report.add("Final Exam 350-450 pts", "LD-PIL-08, LD-PIL-21", "FAIL", f"Unit 14: {final_pts} pts (must be 350-450)")

    # 3) Midterm 100-200 pts (default Unit 7, Primary Pts)
    midterm_pts = primary_pts[6]  # unit 7 → index 6
    if midterm_pts is None:
        report.add("Midterm 100-200 pts", "LD-PIL-21", "FAIL", "Unit 7 Primary Pts is blank")
    elif 100 <= midterm_pts <= 200:
        report.add("Midterm 100-200 pts", "LD-PIL-21", "PASS", f"Unit 7: {midterm_pts} pts")
    else:
        report.add("Midterm 100-200 pts", "LD-PIL-21", "FAIL", f"Unit 7: {midterm_pts} pts (must be 100-200)")

    # 4) Act 1 ≤ 250 pts (Units 1-3 unit pts)
    act_1_total = sum(p for p in unit_pts[0:3] if p is not None)
    report.add(
        "Act 1 ≤ 250 pts",
        "LD-PIL-08",
        "PASS" if act_1_total <= 250 else "FAIL",
        f"Act 1 (Units 1-3): {act_1_total} pts" + ("" if act_1_total <= 250 else " (must be ≤ 250)"),
    )

    # 5) Max 5 units w/ secondary
    secondary_count = sum(1 for s in secondary_assessment if s)
    report.add(
        "Max 5 units w/ secondary",
        "LD-PIL-09",
        "PASS" if secondary_count <= 5 else "FAIL",
        f"{secondary_count} units have secondary assessment" + ("" if secondary_count <= 5 else " (must be ≤ 5)"),
    )

    # 6) No secondary > 15 pts
    max_secondary = max((s for s in secondary_pts if s is not None), default=0)
    report.add(
        "No secondary > 15 pts",
        "LD-PIL-10, LD-GRD-02",
        "PASS" if max_secondary <= 15 else "FAIL",
        f"Max secondary points: {max_secondary}" + ("" if max_secondary <= 15 else " (must be ≤ 15)"),
    )

    # 7) Midterm unit (Unit 7) exam-only — no secondary, no CM
    midterm_clean = not secondary_assessment[6] and not cm_name[6]
    report.add(
        "Midterm unit exam-only",
        "LD-PIL-09",
        "PASS" if midterm_clean else "FAIL",
        "Unit 7 has no secondary or CM" if midterm_clean else f"Unit 7 has extra: secondary='{secondary_assessment[6]}' cm='{cm_name[6]}'",
    )

    # 8) Final unit (Unit 14) exam-only
    final_clean = not secondary_assessment[13] and not cm_name[13]
    report.add(
        "Final unit exam-only",
        "LD-PIL-09",
        "PASS" if final_clean else "FAIL",
        "Unit 14 has no secondary or CM" if final_clean else f"Unit 14 has extra: secondary='{secondary_assessment[13]}' cm='{cm_name[13]}'",
    )

    # 9) ≥ 9 exam-practice primaries (count "Problem Set" in Primary Assessment)
    pset_count = sum(1 for p in primary_assessment if p == "Problem Set")
    report.add(
        "≥ 9 exam-practice primaries",
        "LD-PIL-09",
        "PASS" if pset_count >= 9 else "INFO",
        f"{pset_count} units with Problem Set as primary" + ("" if pset_count >= 9 else " (target: ≥ 9)"),
    )

    # 10) Skill benchmarked in Unit 6 (skill name contains "(B)")
    # The skill benchmark is assigned post-build, not during. During build the
    # cell is expected to be empty or "TBD". Treat that case as INFO so the
    # check doesn't fail loudly during the active build phases.
    unit_6_skill = skill[5]
    has_b = "(B)" in unit_6_skill
    is_unset = unit_6_skill in {"", "TBD", "tbd", "—", "-"}
    if has_b:
        sev = "PASS"
        msg = f"Unit 6 skill: '{unit_6_skill}'"
    elif is_unset:
        sev = "INFO"
        msg = f"Unit 6 skill: '{unit_6_skill}' — skill benchmark assigned post-build (expected during build)"
    else:
        sev = "FAIL"
        msg = f"Unit 6 skill: '{unit_6_skill}' (expected '(B)' marker once benchmark is assigned)"
    report.add(
        "Skill benchmarked in Unit 6",
        "course-template requirement",
        sev,
        msg,
    )


# ---------------------------------------------------------------------------
# LO Map checks (standards-driven, not in spreadsheet)
# ---------------------------------------------------------------------------

def check_lo_map(ws, report: Report) -> None:
    plos: list[tuple[int, str]] = []
    clos: list[tuple[int, str, str]] = []  # (row, number, parent)
    klos: list[tuple[int, str, str]] = []

    for r in range(2, ws.max_row + 1):
        lo_type = _str(ws.cell(r, 1).value)
        number = _str(ws.cell(r, 2).value)
        parent = _str(ws.cell(r, 3).value)
        description = _str(ws.cell(r, 4).value)
        if not lo_type and not description:
            continue
        if lo_type == "PLO":
            plos.append((r, description))
        elif lo_type == "CLO":
            clos.append((r, number, parent))
        elif lo_type == "KLO":
            klos.append((r, number, parent))

    # 1 PLO
    report.add(
        "Exactly 1 PLO",
        "LD-PIL-03",
        "PASS" if len(plos) == 1 else "FAIL",
        f"{len(plos)} PLOs found",
    )

    # ≤ 8 CLOs
    report.add(
        "≤ 8 CLOs",
        "LD-PIL-04",
        "PASS" if len(clos) <= 8 else "FAIL",
        f"{len(clos)} CLOs" + ("" if len(clos) <= 8 else " (must be ≤ 8)"),
    )

    # KLO range 35-55
    if len(klos) >= 35 and len(klos) <= 55:
        sev = "PASS"
    elif len(klos) == 0:
        sev = "INFO"  # template stage; no KLOs yet
    else:
        sev = "FAIL"
    report.add(
        "35-55 KLOs",
        "LD-PIL-20",
        sev,
        f"{len(klos)} KLOs" + (" (template empty)" if len(klos) == 0 else ""),
    )

    # CLOs reference PLO as parent
    bad_clo_parents = [(r, p) for r, _, p in clos if p != "PLO"]
    report.add(
        "Every CLO has Parent=PLO",
        "LD-PIL-04",
        "PASS" if not bad_clo_parents else "FAIL",
        "All CLOs reference PLO" if not bad_clo_parents else f"Bad CLO parents: {bad_clo_parents}",
    )

    # KLOs reference a CLO as parent (CLO-N format)
    bad_klo_parents = [(r, p) for r, _, p in klos if not re.match(r"^CLO-\d+", p)]
    report.add(
        "Every KLO references a CLO",
        "LD-PIL-20",
        "PASS" if not bad_klo_parents else "FAIL",
        "All KLOs reference a CLO" if not bad_klo_parents else f"Bad KLO parents: {bad_klo_parents}",
    )


# ---------------------------------------------------------------------------
# Course Narrative checks (six sections populated)
# ---------------------------------------------------------------------------

REQUIRED_SECTIONS = [
    "Course Opinion",
    "Act 1 Narrative",
    "Act 2 Narrative",
    "Act 3 Narrative",
    "What to Avoid",
    "What Success Looks Like",
]


def check_course_narrative(ws, report: Report) -> None:
    """Check Course Narrative sections.

    Handles two layout patterns:
    1. Chemistry-exemplar style: header + content in the same cell (e.g.,
       'Course Opinion: You\\'re going to learn...').
    2. Canonical-template style: header in one cell, guidance in next cell,
       placeholder/content in subsequent cells.

    A section counts as populated when the cell starting with the header has
    substantial content beyond the header itself, OR a subsequent non-guidance
    non-placeholder cell contains substantial content.
    """
    rows = []
    for r in range(1, ws.max_row + 1):
        v = _str(ws.cell(r, 1).value)
        if v:
            rows.append((r, v))

    CONTENT_THRESHOLD_CHARS = 80  # cells with more than this much non-header content count as populated

    section_state = {name: {"found": False, "has_content": False, "header_inline_chars": 0} for name in REQUIRED_SECTIONS}
    current = None

    for _, v in rows:
        # Check if this row begins a section header
        matched_section = None
        for name in REQUIRED_SECTIONS:
            if v.startswith(name):
                matched_section = name
                break

        if matched_section:
            section_state[matched_section]["found"] = True
            current = matched_section
            # Pattern 1: header + content in the same cell
            # Strip the header name (and optional separator like ":" or "—") and see what remains
            remainder = v[len(matched_section):].lstrip(":–—- \t")
            if len(remainder) >= CONTENT_THRESHOLD_CHARS:
                section_state[matched_section]["has_content"] = True
            continue

        # Not a header — check if this row is content for the current section
        if current and section_state[current]["found"]:
            is_guidance = v.lower().startswith("guidance:")
            is_placeholder = v.startswith("[") and v.endswith("]")
            if not is_guidance and not is_placeholder and len(v) >= CONTENT_THRESHOLD_CHARS:
                section_state[current]["has_content"] = True

    for name in REQUIRED_SECTIONS:
        st = section_state[name]
        if not st["found"]:
            report.add(f"Course Narrative: {name} present", "design-lock", "FAIL", f"Section '{name}' not found")
        elif not st["has_content"]:
            report.add(f"Course Narrative: {name} populated", "design-lock", "INFO", f"Section '{name}' has only header / guidance / placeholder")
        else:
            report.add(f"Course Narrative: {name} populated", "design-lock", "PASS", f"Section '{name}' has substantial content")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_checks(xlsx_path: Path) -> Report:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Course plan xlsx not found: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    report = Report(artifact=str(xlsx_path))

    if "Assessment Map" in wb.sheetnames:
        check_assessment_map(wb["Assessment Map"], report)
    else:
        report.add("Assessment Map sheet present", "structural", "FAIL", "No 'Assessment Map' sheet found")

    if "LO Map" in wb.sheetnames:
        check_lo_map(wb["LO Map"], report)
    else:
        report.add("LO Map sheet present", "structural", "FAIL", "No 'LO Map' sheet found")

    if "Course Narrative" in wb.sheetnames:
        check_course_narrative(wb["Course Narrative"], report)
    else:
        report.add("Course Narrative sheet present", "design-lock", "FAIL", "No 'Course Narrative' sheet found")

    # Cross-sheet check: every CLO listed as covered in any unit must have all
    # its KLOs assigned to at least one unit's KLOs Covered field.
    if "Assessment Map" in wb.sheetnames and "LO Map" in wb.sheetnames:
        check_klo_distribution(wb["Assessment Map"], wb["LO Map"], report)

    return report


# ---------------------------------------------------------------------------
# Cross-sheet KLO distribution check
# ---------------------------------------------------------------------------

def check_klo_distribution(ws_assess, ws_lo, report: Report) -> None:
    """For every CLO covered in any unit, verify all its KLOs are assigned to a unit.

    Surfaces the Phase 4 oversight where a CLO is listed as covered in a unit
    but only some of its KLOs are listed in the unit's KLOs Covered field, leaving
    other KLOs of that CLO orphaned across the schedule.
    """
    # Build CLO → set of KLO numbers from the LO Map
    clo_to_klos: dict[str, set[str]] = {}
    for r in range(2, ws_lo.max_row + 1):
        lo_type = _str(ws_lo.cell(r, 1).value)
        number = _str(ws_lo.cell(r, 2).value)
        parent = _str(ws_lo.cell(r, 3).value)
        if lo_type == "KLO" and parent.startswith("CLO-") and number:
            # Number may be "1.0" — keep as integer-string for matching
            klo_num = number.split(".")[0] if "." in number else number
            clo_key = parent.replace("CLO-", "").split(".")[0]
            clo_to_klos.setdefault(clo_key, set()).add(klo_num)

    # Build CLO → set of units where it's covered, and union of KLOs covered there
    clo_to_units_covered: dict[str, set[int]] = {}
    klos_covered_by_any_unit: set[tuple[str, str]] = set()  # (clo, klo) pairs covered somewhere

    for r in ASSESSMENT_DATA_ROWS:
        unit = _num(ws_assess.cell(r, COL["Unit"]).value)
        clos_covered = _str(ws_assess.cell(r, COL["CLOs Covered"]).value)
        klos_covered_str = _str(ws_assess.cell(r, COL["KLOs Covered"]).value)

        # Parse CLOs Covered: e.g., "CLO-1,2,3" or "CLO-1, CLO-2" or "All"
        clos_in_unit: list[str] = []
        if clos_covered.lower() in ("all", "all clos"):
            clos_in_unit = list(clo_to_klos.keys())
        else:
            cleaned = clos_covered.replace("CLO-", "").replace("CLO ", "")
            for part in cleaned.replace(";", ",").split(","):
                p = part.strip().split(".")[0]
                if p and p.isdigit():
                    clos_in_unit.append(p)

        for clo in clos_in_unit:
            clo_to_units_covered.setdefault(clo, set()).add(int(unit) if unit else 0)

        # Parse KLOs Covered: e.g., "1, 2, 3" or "12-18" or "KLO-1, KLO-3" or "1-11"
        klo_nums_in_unit: set[str] = set()
        cleaned_klos = klos_covered_str.replace("KLO-", "").replace("KLO ", "")
        for part in cleaned_klos.replace(";", ",").split(","):
            p = part.strip()
            if not p:
                continue
            # Handle ranges like "12-18" or "1–11"
            if "-" in p or "–" in p:
                sep = "-" if "-" in p else "–"
                try:
                    start_s, end_s = p.split(sep)
                    start, end = int(start_s.strip()), int(end_s.strip())
                    for n in range(start, end + 1):
                        klo_nums_in_unit.add(str(n))
                except ValueError:
                    continue
            elif p.split(".")[0].isdigit():
                klo_nums_in_unit.add(p.split(".")[0])

        # Mark every (clo, klo) pair in this unit as covered
        for clo in clos_in_unit:
            for klo in klo_nums_in_unit:
                klos_covered_by_any_unit.add((clo, klo))

    # Now check: for each CLO covered in any unit, are all its KLOs covered somewhere?
    orphaned: list[tuple[str, set[str]]] = []
    for clo, units in clo_to_units_covered.items():
        if not units:
            continue
        expected_klos = clo_to_klos.get(clo, set())
        actually_covered = {klo for (c, klo) in klos_covered_by_any_unit if c == clo}
        missing = expected_klos - actually_covered
        if missing:
            orphaned.append((clo, missing))

    if not orphaned:
        report.add(
            "All KLOs distributed to a covering unit",
            "LD-PIL-20",
            "PASS",
            f"Every CLO's KLOs appear in at least one unit's KLOs Covered field",
        )
    else:
        details = "; ".join(
            f"CLO-{clo}: missing KLO(s) {sorted(missing, key=lambda x: int(x) if x.isdigit() else 9999)}"
            for clo, missing in orphaned
        )
        report.add(
            "All KLOs distributed to a covering unit",
            "LD-PIL-20",
            "FAIL",
            f"Found CLOs with KLOs not assigned to any unit: {details}",
        )


def render_report(report: Report) -> str:
    lines = [
        f"Course Plan Check — {report.artifact}",
        f"  PASS: {report.pass_count}   FAIL: {report.fail_count}   INFO: {report.info_count}",
        "",
    ]
    for r in report.results:
        marker = {"PASS": "✓", "FAIL": "✗", "INFO": "·"}.get(r.severity, "?")
        lines.append(f"  [{marker} {r.severity}] {r.name}")
        lines.append(f"      Standard: {r.standard}")
        lines.append(f"      {r.message}")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="D1 mechanical validation for the course plan xlsx.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--course", type=Path, help="Course working directory (looks for design/course-plan-exam.xlsx)")
    group.add_argument("--xlsx", type=Path, help="Direct path to the course plan xlsx")
    parser.add_argument("--report", type=Path, help="Write JSON report to this path")
    parser.add_argument("--strict", action="store_true", help="Exit 1 if any FAIL")
    args = parser.parse_args()

    if args.course:
        xlsx_path = args.course / "design" / "course-plan-exam.xlsx"
    else:
        xlsx_path = args.xlsx

    try:
        report = run_checks(xlsx_path)
    except FileNotFoundError as e:
        print(f"[check_course_plan] ERROR: {e}", file=sys.stderr)
        return 2

    print(render_report(report))

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report.to_dict(), indent=2), encoding="utf-8")
        print(f"\nJSON report: {args.report}", file=sys.stderr)

    if args.strict and report.fail_count > 0:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
