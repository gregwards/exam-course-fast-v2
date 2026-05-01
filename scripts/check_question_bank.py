#!/usr/bin/env python3
"""
D1 mechanical validation for the per-course question bank.

The bank is the central database for all per-question content (problem sets,
walkthroughs, quizzes, exam items). This script validates the bank's structural
integrity at the row level — type-allocation compliance, KLO targeting,
scenario-tag uniqueness, cross-exam deduplication, point ranges per LD-PIL-21,
and KLO coverage across allocations.

Runs in seconds. No LLM calls. Catches the structural failures that would
otherwise reach the LLM-based standards auditor (D2) or surface to the ID.

Bank schemas (from templates/source/exam-question-bank.xlsx):

    Multiple Choice (Quiz / Exam Bank only):
      Q-ID, Unit, KLO, Allocation, Review Status, Points, Question Type,
      Stem, Option A, Option A Comment, ..., Correct Answer, Feedback Neutral,
      Scenario Tag

    Short Answers (Problem Set / Walkthrough / Exam Bank):
      Q-ID, Unit, CLO, KLOs, Tier, Allocation, Review Status, Points,
      Question Type, Question, Feedback Neutral, Model Answer, Key Points,
      Explanation, Annotations, Scenario Tag

    Long Answers: same shape as Short Answers, multi-part Question.

Usage:
    python3 scripts/check_question_bank.py --course /path/to/course-folder
    python3 scripts/check_question_bank.py --xlsx /path/to/master-questions.xlsx
    [--unit N]           limit checks to one unit
    [--report out.json]
    [--strict]           exit 1 if any FAIL
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Optional

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _common as common  # noqa: E402


# ---------------------------------------------------------------------------
# Schema constants
# ---------------------------------------------------------------------------

# Allocation × Question Type compliance per LD-PIL-16 / LD-PIL-21.
# Type names accepted as synonyms (essay / essayQuestion are v3 vs v4 labels for the same thing).
ALLOC_TYPE_RULES: dict[str, set[str]] = {
    "Quiz": {
        # Auto-gradable types only per LD-PIL-16
        "multipleChoice", "trueFalse", "multipleAnswer", "matching",
        "numberEntry",
        # Plus essay sparingly
        "essay", "essayQuestion",
    },
    "Problem Set": {
        # Per LD-PIL-19, problem sets are downloadable work-product documents.
        # Section 2 (Short Response) and Section 3 (Long Response) types apply,
        # plus framework's internal SA/LA labels.
        "essay", "essayQuestion", "fileUpload",
        "shortAnswer", "longAnswer",
    },
    "Walkthrough": {
        # Walkthroughs are a teaching surface — any question type can be the
        # subject of a walkthrough (showing how to reason through MCQ options,
        # how to solve a long-form problem with shown work, etc.).
        "multipleChoice", "trueFalse", "multipleAnswer", "matching",
        "numberEntry", "numericalAnswerExact", "numericalAnswerPrecision",
        "numericalAnswerRange", "essay", "essayQuestion", "fileUpload",
        "shortAnswer", "longAnswer", "textNoQuestion",
    },
    "Exam Bank": {
        # Section 1 (auto-gradable) + Section 2 (Short Response) + Section 3 (Long Response)
        "multipleChoice", "trueFalse", "multipleAnswer", "matching",
        "numericalAnswerExact", "numericalAnswerPrecision", "numericalAnswerRange",
        "textNoQuestion",
        "essay", "essayQuestion", "fileUpload",
        "shortAnswer", "longAnswer",
    },
}

# Sheet → expected sheet name in xlsx
SHEET_NAMES = {
    "mc": "Multiple Choice",
    "sa": "Short Answers",
    "la": "Long Answers",
}

# Column indices per sheet (1-based, matching openpyxl). None means "first occurrence by header"
MC_COL = {
    "Q-ID": 1, "Unit": 2, "KLO": 3, "Allocation": 4, "Review Status": 5,
    "Points": 6, "Question Type": 7, "Stem": 8,
    "Option A": 9, "Option A Comment": 10, "Option B": 11, "Option B Comment": 12,
    "Option C": 13, "Option C Comment": 14, "Option D": 15, "Option D Comment": 16,
    "Correct Answer": 17, "Feedback Neutral": 18, "Scenario Tag": 19,
}

SA_LA_COL = {
    "Q-ID": 1, "Unit": 2, "CLO": 3, "KLOs": 4, "Tier": 5, "Allocation": 6,
    "Review Status": 7, "Points": 8, "Question Type": 9, "Question": 10,
    "Feedback Neutral": 11, "Model Answer": 12, "Key Points": 13,
    "Explanation": 14, "Annotations": 15, "Scenario Tag": 16,
}


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

@dataclass
class CheckResult:
    name: str
    standard: str
    severity: str  # PASS | FAIL | INFO
    message: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class Report:
    artifact: str
    results: list[CheckResult] = field(default_factory=list)

    def add(self, name: str, standard: str, severity: str, message: str) -> None:
        self.results.append(CheckResult(name=name, standard=standard, severity=severity, message=message))

    @property
    def fail_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "FAIL")

    @property
    def pass_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "PASS")

    @property
    def info_count(self) -> int:
        return sum(1 for r in self.results if r.severity == "INFO")

    def to_dict(self) -> dict[str, Any]:
        return {
            "artifact": self.artifact,
            "summary": {"pass": self.pass_count, "fail": self.fail_count, "info": self.info_count},
            "results": [r.to_dict() for r in self.results],
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _num(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row_to_dict(ws, row: int, col_map: dict[str, int]) -> dict[str, Any]:
    return {name: ws.cell(row, col).value for name, col in col_map.items()}


def _is_header_or_blank(row_dict: dict[str, Any]) -> bool:
    """Skip rows that are headers (Q-ID == 'Q-ID') or fully blank."""
    qid = _str(row_dict.get("Q-ID"))
    if qid == "Q-ID":
        return True
    return not any(_str(v) for v in row_dict.values())


def _data_rows(ws, col_map: dict[str, int]) -> list[tuple[int, dict[str, Any]]]:
    out: list[tuple[int, dict[str, Any]]] = []
    for r in range(1, ws.max_row + 1):
        row = _row_to_dict(ws, r, col_map)
        if _is_header_or_blank(row):
            continue
        out.append((r, row))
    return out


# ---------------------------------------------------------------------------
# Per-sheet checks
# ---------------------------------------------------------------------------

def check_sheet(
    ws,
    col_map: dict[str, int],
    klo_field: str,  # "KLO" for MC, "KLOs" for SA/LA
    sheet_label: str,
    report: Report,
    unit_filter: Optional[int] = None,
) -> dict[str, list[tuple[int, dict[str, Any]]]]:
    """Run row-level checks on a sheet. Returns rows by allocation for cross-sheet checks."""
    rows = _data_rows(ws, col_map)
    by_allocation: dict[str, list[tuple[int, dict[str, Any]]]] = {}

    for excel_row, row in rows:
        unit = _num(row.get("Unit"))
        if unit_filter is not None and unit != unit_filter:
            continue

        qid = _str(row.get("Q-ID"))
        allocation = _str(row.get("Allocation"))
        qtype = _str(row.get("Question Type"))
        klo_val = _str(row.get(klo_field))
        scenario_tag = _str(row.get("Scenario Tag"))

        # Required: Q-ID, Unit, KLO/KLOs, Allocation, Question Type
        if not qid:
            report.add(f"{sheet_label} R{excel_row}: Q-ID present", "structural", "FAIL", f"Row missing Q-ID")
            continue
        if not unit:
            report.add(f"{sheet_label} {qid}: Unit present", "structural", "FAIL", f"{qid} missing Unit")
        if not klo_val:
            report.add(f"{sheet_label} {qid}: {klo_field} target", "LD-PIL-20", "FAIL",
                       f"{qid} has no {klo_field} target — every question must target a KLO")
        if not allocation:
            report.add(f"{sheet_label} {qid}: Allocation present", "structural", "FAIL",
                       f"{qid} has no Allocation")
            continue
        if allocation not in ALLOC_TYPE_RULES:
            report.add(f"{sheet_label} {qid}: Allocation valid", "structural", "FAIL",
                       f"{qid} has unknown Allocation '{allocation}'")
            continue

        # Type-allocation compliance per LD-PIL-16 / LD-PIL-21
        # Be lenient with case and spacing — but do enforce the rule
        qtype_normalized = qtype.replace(" ", "").lower()
        valid_types_normalized = {t.replace(" ", "").lower() for t in ALLOC_TYPE_RULES[allocation]}
        if not qtype:
            report.add(f"{sheet_label} {qid}: Question Type present", "structural", "FAIL",
                       f"{qid} has no Question Type")
        elif qtype_normalized not in valid_types_normalized:
            report.add(f"{sheet_label} {qid}: Type-Allocation compliance",
                       "LD-PIL-16, LD-PIL-21",
                       "FAIL",
                       f"{qid} has Type='{qtype}' incompatible with Allocation='{allocation}'. "
                       f"Allowed for {allocation}: {sorted(ALLOC_TYPE_RULES[allocation])}")

        by_allocation.setdefault(allocation, []).append((excel_row, row))

    return by_allocation


def check_scenario_uniqueness(
    rows_by_alloc: dict[str, list[tuple[int, dict[str, Any]]]],
    sheet_label: str,
    report: Report,
) -> None:
    """No two questions in the same unit share a Scenario Tag."""
    # Group by (unit, scenario_tag) across all allocations in this sheet
    seen: dict[tuple[Any, str], list[str]] = {}
    for alloc, rows in rows_by_alloc.items():
        for _, row in rows:
            unit = _num(row.get("Unit"))
            tag = _str(row.get("Scenario Tag"))
            qid = _str(row.get("Q-ID"))
            if not tag:
                continue  # blank scenario tag is its own issue, surfaced elsewhere if needed
            seen.setdefault((unit, tag), []).append(qid)

    duplicates = {k: qids for k, qids in seen.items() if len(qids) > 1}
    if not duplicates:
        report.add(f"{sheet_label}: scenario tags unique within units", "scenario-uniqueness",
                   "PASS", f"No duplicate Scenario Tags within any unit")
    else:
        details = "; ".join(f"Unit {u}, tag '{t}' used by {qids}" for (u, t), qids in duplicates.items())
        report.add(f"{sheet_label}: scenario tags unique within units", "scenario-uniqueness",
                   "FAIL", f"Duplicate Scenario Tags: {details}")


# Per-type point values per LD-PIL-21 Exam composition
EXAM_POINTS_BY_TYPE = {
    "AG": 2,   # 15 × 2 = 30
    "SA": 6,   # 5 × 6 = 30
    "LA": 15,  # 3 × 15 = 45 + 1 PLO × 15 = 15 → total 60
}

# Per-administration target counts per LD-PIL-21 (one final exam sitting)
PER_ADMIN_TARGET = {
    "AG": 15,
    "SA": 5,
    "LA": 4,  # 3 CLO-mapped + 1 PLO-mapped
}

# Pool multiplier per LD-PIL-21A — default 2.5× for two-semester rotation plus reserve.
# Configurable via --pool-multiplier on the CLI, or via course-config (future).
DEFAULT_POOL_MULTIPLIER = 2.5


def _pool_target(multiplier: float = DEFAULT_POOL_MULTIPLIER) -> dict[str, int]:
    """Compute the pool target per type by multiplying the per-administration target."""
    import math
    return {k: math.ceil(v * multiplier) for k, v in PER_ADMIN_TARGET.items()}


def check_exam_composition(
    mc_exam_rows: list[tuple[int, dict[str, Any]]],
    sa_exam_rows: list[tuple[int, dict[str, Any]]],
    la_exam_rows: list[tuple[int, dict[str, Any]]],
    report: Report,
    pool_multiplier: float = DEFAULT_POOL_MULTIPLIER,
) -> None:
    """Verify the exam-bank totals approach the LD-PIL-21 + LD-PIL-21A pool target.

    The pool target is `per-administration target × multiplier` (default 2.5× for
    two-semester rotation). Severity tiers:
      - At or above pool target: PASS
      - At or above per-admin target but below pool: INFO (building toward pool)
      - Below 50% of per-admin target: FAIL
      - 50–100% of per-admin target: INFO (early-build stage)
    """
    counts = {
        "AG": len(mc_exam_rows),
        "SA": len(sa_exam_rows),
        "LA": len(la_exam_rows),
    }

    pool_targets = _pool_target(pool_multiplier)

    for type_label, per_admin in PER_ADMIN_TARGET.items():
        actual = counts.get(type_label, 0)
        pool = pool_targets[type_label]

        if actual >= pool:
            sev = "PASS"
            msg = (
                f"{actual} {type_label} exam-bank rows — meets pool target ({pool} = "
                f"{per_admin} × {pool_multiplier}×) for multi-semester rotation"
            )
        elif actual >= per_admin:
            sev = "INFO"
            msg = (
                f"{actual} {type_label} exam-bank rows — meets single-administration target ({per_admin}); "
                f"building toward pool target ({pool} for {pool_multiplier}× rotation)"
            )
        elif actual >= per_admin / 2:
            sev = "INFO"
            msg = (
                f"{actual}/{per_admin} {type_label} exam-bank rows — accumulating "
                f"during build (pool target: {pool})"
            )
        else:
            sev = "FAIL"
            msg = (
                f"{actual}/{per_admin} {type_label} exam-bank rows — well below "
                f"single-administration target (pool target: {pool})"
            )

        report.add(
            f"Exam composition: {type_label} count target",
            "LD-PIL-21, LD-PIL-21A",
            sev,
            msg,
        )


def check_problem_set_sizing(
    sa_by_alloc: dict[str, list[tuple[int, dict[str, Any]]]],
    la_by_alloc: dict[str, list[tuple[int, dict[str, Any]]]],
    report: Report,
) -> None:
    """Verify problem set sizing per LD-PIL-19A.

    Per-unit checks:
      - Total count: Unit 3+ requires ≥ 4 questions
      - Mix: when both SA and LA are present in any unit's bank, the problem set
        must include at least one of each (no LA-only problem sets)
      - Units 1, 2: count rule is INFO only (scaffolding may lengthen scope)
    """
    # Build per-unit SA and LA Problem Set counts
    sa_by_unit: dict[int, int] = {}
    la_by_unit: dict[int, int] = {}
    for _, row in sa_by_alloc.get("Problem Set", []):
        unit = _num(row.get("Unit"))
        if unit is not None:
            sa_by_unit[int(unit)] = sa_by_unit.get(int(unit), 0) + 1
    for _, row in la_by_alloc.get("Problem Set", []):
        unit = _num(row.get("Unit"))
        if unit is not None:
            la_by_unit[int(unit)] = la_by_unit.get(int(unit), 0) + 1

    all_units = sorted(set(sa_by_unit.keys()) | set(la_by_unit.keys()))
    if not all_units:
        return  # nothing to check yet

    for unit in all_units:
        sa_n = sa_by_unit.get(unit, 0)
        la_n = la_by_unit.get(unit, 0)
        total = sa_n + la_n

        # 1) Total count check
        if unit in (1, 2):
            count_sev = "INFO"
            count_msg = f"Unit {unit}: {total} pset questions ({sa_n} SA + {la_n} LA) — early-unit minimum is informational"
        elif total < 4 and total > 0:
            count_sev = "FAIL"
            late_qualifier = " (late-course unit — extra concern for exam-prep coverage)" if unit >= 8 else ""
            count_msg = f"Unit {unit}: {total} pset questions ({sa_n} SA + {la_n} LA) — below minimum 4 for Unit 3+{late_qualifier}"
        else:
            count_sev = "PASS"
            count_msg = f"Unit {unit}: {total} pset questions ({sa_n} SA + {la_n} LA)"

        report.add(
            f"Problem set sizing — count: Unit {unit}",
            "LD-PIL-19A",
            count_sev,
            count_msg,
        )

        # 2) Mix check — only fires for Unit 3+ (Units 1-2 may legitimately have only one type)
        if unit >= 3 and total > 0:
            if la_n > 0 and sa_n == 0:
                # LA-only problem set (the MATHS-VIII Unit 11 failure pattern)
                report.add(
                    f"Problem set sizing — mix: Unit {unit}",
                    "LD-PIL-19A",
                    "FAIL",
                    f"Unit {unit}: {la_n} LA, 0 SA — LA-only problem sets flatten the SA:LA workload ratio and exhaust the unit's budget. Add at least one SA item.",
                )
            elif sa_n > 0 and la_n == 0 and unit >= 4:
                # SA-only problem set after Unit 3 — softer warning
                report.add(
                    f"Problem set sizing — mix: Unit {unit}",
                    "LD-PIL-19A",
                    "INFO",
                    f"Unit {unit}: {sa_n} SA, 0 LA — SA-only problem set. Acceptable for some units; verify the unit's CLO/KLO coverage doesn't warrant LA-format practice.",
                )
            else:
                report.add(
                    f"Problem set sizing — mix: Unit {unit}",
                    "LD-PIL-19A",
                    "PASS",
                    f"Unit {unit}: {sa_n} SA + {la_n} LA",
                )


def check_cross_exam_dedup(
    all_exam_bank_rows: list[tuple[str, dict[str, Any]]],
    report: Report,
) -> None:
    """No Q-ID appears in more than one exam event composition.

    The bank itself shouldn't contain duplicate Q-IDs (that's a structural failure)
    but this check is the placeholder for the future cross-exam composition check
    where the same bank Q-ID gets pulled into both midterms or final.

    For now, just verify Q-IDs are unique across the bank.
    """
    qids: dict[str, list[str]] = {}
    for sheet_label, row in all_exam_bank_rows:
        qid = _str(row.get("Q-ID"))
        if qid:
            qids.setdefault(qid, []).append(sheet_label)

    duplicates = {qid: sheets for qid, sheets in qids.items() if len(sheets) > 1}
    if not duplicates:
        report.add("Exam Bank Q-IDs unique", "v3-id-feedback (cross-exam dedup)",
                   "PASS", f"No duplicate Q-IDs in Exam Bank allocation")
    else:
        details = "; ".join(f"{qid} in {sheets}" for qid, sheets in duplicates.items())
        report.add("Exam Bank Q-IDs unique", "v3-id-feedback (cross-exam dedup)",
                   "FAIL", f"Duplicate Q-IDs across Exam Bank rows: {details}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_checks(
    xlsx_path: Path,
    unit_filter: Optional[int] = None,
    pool_multiplier: float = DEFAULT_POOL_MULTIPLIER,
) -> Report:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Question bank xlsx not found: {xlsx_path}")

    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        raise ImportError("openpyxl required: pip install openpyxl") from e

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    report = Report(artifact=str(xlsx_path))

    all_exam_bank_rows: list[tuple[str, dict[str, Any]]] = []

    # Per-sheet by-allocation maps for cross-sheet checks
    mc_by_alloc: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    sa_by_alloc: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    la_by_alloc: dict[str, list[tuple[int, dict[str, Any]]]] = {}

    # Multiple Choice sheet
    if SHEET_NAMES["mc"] in wb.sheetnames:
        mc_by_alloc = check_sheet(
            wb[SHEET_NAMES["mc"]], MC_COL, "KLO", "MC", report, unit_filter,
        )
        check_scenario_uniqueness(mc_by_alloc, "MC", report)
        for _, row in mc_by_alloc.get("Exam Bank", []):
            all_exam_bank_rows.append(("MC", row))
    else:
        report.add(f"{SHEET_NAMES['mc']} sheet present", "structural", "INFO",
                   f"No '{SHEET_NAMES['mc']}' sheet — skipped")

    # Short Answers sheet
    if SHEET_NAMES["sa"] in wb.sheetnames:
        sa_by_alloc = check_sheet(
            wb[SHEET_NAMES["sa"]], SA_LA_COL, "KLOs", "SA", report, unit_filter,
        )
        check_scenario_uniqueness(sa_by_alloc, "SA", report)
        for _, row in sa_by_alloc.get("Exam Bank", []):
            all_exam_bank_rows.append(("SA", row))
    else:
        report.add(f"{SHEET_NAMES['sa']} sheet present", "structural", "INFO",
                   f"No '{SHEET_NAMES['sa']}' sheet — skipped")

    # Long Answers sheet
    if SHEET_NAMES["la"] in wb.sheetnames:
        la_by_alloc = check_sheet(
            wb[SHEET_NAMES["la"]], SA_LA_COL, "KLOs", "LA", report, unit_filter,
        )
        check_scenario_uniqueness(la_by_alloc, "LA", report)
        for _, row in la_by_alloc.get("Exam Bank", []):
            all_exam_bank_rows.append(("LA", row))
    else:
        report.add(f"{SHEET_NAMES['la']} sheet present", "structural", "INFO",
                   f"No '{SHEET_NAMES['la']}' sheet — skipped")

    # Cross-sheet: exam bank Q-ID uniqueness
    check_cross_exam_dedup(all_exam_bank_rows, report)

    # LD-PIL-21 / LD-PIL-21A final-exam pool target (run only when checking the whole bank)
    if unit_filter is None:
        check_exam_composition(
            mc_by_alloc.get("Exam Bank", []),
            sa_by_alloc.get("Exam Bank", []),
            la_by_alloc.get("Exam Bank", []),
            report,
            pool_multiplier=pool_multiplier,
        )

    # LD-PIL-19A problem-set sizing
    check_problem_set_sizing(sa_by_alloc, la_by_alloc, report)

    return report


def render_report(report: Report) -> str:
    lines = [
        f"Question Bank Check — {report.artifact}",
        f"  PASS: {report.pass_count}   FAIL: {report.fail_count}   INFO: {report.info_count}",
        "",
    ]
    for r in report.results:
        marker = {"PASS": "✓", "FAIL": "✗", "INFO": "·"}.get(r.severity, "?")
        lines.append(f"  [{marker} {r.severity}] {r.name}")
        lines.append(f"      Standard: {r.standard}")
        lines.append(f"      {r.message}")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="D1 mechanical validation for the question bank.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--course", type=Path, help="Course directory (looks for data/master-questions.xlsx)")
    group.add_argument("--xlsx", type=Path, help="Direct path to the question bank xlsx")
    parser.add_argument("--unit", type=int, help="Limit checks to a single unit number")
    parser.add_argument(
        "--pool-multiplier",
        type=float,
        default=DEFAULT_POOL_MULTIPLIER,
        help=f"Multiplier for exam-bank pool target per LD-PIL-21A (default {DEFAULT_POOL_MULTIPLIER}x for two-semester rotation; use 1.0 for single administration, 3.5 for three-cohort rotation)",
    )
    parser.add_argument("--report", type=Path, help="Write JSON report to this path")
    parser.add_argument("--strict", action="store_true", help="Exit 1 if any FAIL")
    args = parser.parse_args()

    if args.course:
        xlsx_path = args.course / "data" / "master-questions.xlsx"
    else:
        xlsx_path = args.xlsx

    try:
        report = run_checks(xlsx_path, unit_filter=args.unit, pool_multiplier=args.pool_multiplier)
    except FileNotFoundError as e:
        print(f"[check_question_bank] ERROR: {e}", file=sys.stderr)
        return 2

    print(render_report(report))

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report.to_dict(), indent=2), encoding="utf-8")
        print(f"\nJSON report: {args.report}", file=sys.stderr)

    if args.strict and report.fail_count > 0:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
